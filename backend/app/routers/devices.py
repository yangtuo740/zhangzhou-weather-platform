from fastapi import APIRouter, Depends, HTTPException, Query
from sqlmodel import Session
from ..database import get_session
from ..schemas import DeviceCreate, DeviceUpdate, DeviceOut, DeviceListResponse, ExportRequest, AuditLogOut, AuditLogListResponse
from ..auth import get_current_user
from ..models import User
from ..services.device_service import (
    get_devices, get_device_by_id, create_device, update_device, delete_device, get_audit_logs
)
from ..services.user_service import add_log
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
from fastapi.responses import StreamingResponse

router = APIRouter(prefix="/api/devices", tags=["devices"])

DEVICE_FIELDS = {
    "id": "ID", "cert_no": "证书编号", "send_unit": "送检单位", "device_name": "计量器具名称",
    "model_spec": "型号/规格", "equip_code": "装备编码", "manufacturer": "制造单位",
    "approver": "批准人", "reviewer": "核验员", "calibrator": "校准员",
    "calib_date": "校准日期", "seal": "印章", "status": "状态", "remarks": "备注",
    "created_at": "创建时间", "updated_at": "更新时间"
}

STATUS_MAP = {"active": "在校准期内", "expiring": "即将到期", "expired": "已过期"}

@router.get("")
def list_devices(
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=100),
    keyword: str = Query(""), status: str = Query(""),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    skip = (page - 1) * page_size
    items, total = get_devices(session, skip, page_size, keyword, status)
    return {"items": [DeviceOut.model_validate(d) for d in items], "total": total, "page": page, "page_size": page_size}

@router.get("/{device_id}", response_model=DeviceOut)
def get_device(device_id: int, session: Session = Depends(get_session),
               current_user: User = Depends(get_current_user)):
    device = get_device_by_id(session, device_id)
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")
    return device

@router.post("", response_model=DeviceOut)
def add_device(data: DeviceCreate, session: Session = Depends(get_session),
               current_user: User = Depends(get_current_user)):
    return create_device(session, data, current_user.id, current_user.username)

@router.put("/{device_id}", response_model=DeviceOut)
def edit_device(device_id: int, data: DeviceUpdate, session: Session = Depends(get_session),
                current_user: User = Depends(get_current_user)):
    device = update_device(session, device_id, data, current_user.id, current_user.username)
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")
    return device

@router.delete("/{device_id}")
def remove_device(device_id: int, session: Session = Depends(get_session),
                  current_user: User = Depends(get_current_user)):
    ok = delete_device(session, device_id, current_user.id, current_user.username)
    if not ok:
        raise HTTPException(status_code=404, detail="设备不存在")
    return {"message": "已删除"}

@router.post("/export")
def export_devices(req: ExportRequest, session: Session = Depends(get_session),
                   current_user: User = Depends(get_current_user)):
    items, _ = get_devices(session, export_all=True)
    if req.device_ids:
        ids = set(req.device_ids)
        items = [d for d in items if d.id in ids]

    fields = req.fields or ["cert_no", "send_unit", "device_name", "model_spec", "equip_code",
                             "manufacturer", "approver", "reviewer", "calibrator", "calib_date", "seal", "status"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "设备清单"
    header_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for col_idx, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=DEVICE_FIELDS.get(field, field))
        cell.font = header_font
        cell.border = thin_border

    for row_idx, device in enumerate(items, 2):
        for col_idx, field in enumerate(fields, 1):
            val = getattr(device, field, "")
            if field == "status":
                val = STATUS_MAP.get(val, val)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    ws.column_dimensions["A"].width = 12
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    add_log(session, current_user.id, current_user.username, "export", f"导出 {len(items)} 条设备数据")
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=devices_export.xlsx"})

# ---- Audit Logs ----
@router.get("/logs/list")
def list_audit_logs(
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=100),
    keyword: str = Query(""),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    skip = (page - 1) * page_size
    items, total = get_audit_logs(session, skip, page_size, keyword)
    return {"items": [AuditLogOut.model_validate(l) for l in items], "total": total, "page": page, "page_size": page_size}
